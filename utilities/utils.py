import inspect
import softest
import logging
import csv
from openpyxl import Workbook, load_workbook


class Utils(softest.TestCase):

    def UtilitiesList(self, list, value):
        for stop in list:
            print("The text is: " + stop.text)
            self.soft_assert(self.assertEqual, stop.text, value)
            if stop.text == value:
                print("test pass")
            else:
                print("test fail")

        self.assert_all()

    """this the custom logger"""

    def custom_log(logLevel=logging.DEBUG):
        """set class/method name from where its called"""
        logger_name = inspect.stack()[1][3]
        """create logger and set level"""
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)
        fh = logging.FileHandler("C:\\Users\\Hp\\PycharmProjects\\TestFrameworkMentor\\reports\\exceptionlog.log")
        """create formatter - how you want your log to be formatted"""
        formatter = logging.Formatter('%(asctime)s- %(levelname)s- %(name)s: %(message)s')
        """add formatter to file handler"""
        fh.setFormatter(formatter)
        logger.addHandler(fh)
        return logger

    """create utility for read data from excel"""

    def read_data_from_excel(file_name, sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column
        for r in range(2, row_ct + 1):
            row_data = []
            for c in range(1, col_ct + 1):
                row_data.append(sh.cell(row=r, column=c).value)

            datalist.append(row_data)
        return datalist

    """create utility for read data from csv"""

    def read_data_from_csv(filename):
        # create an empty list that will read data from csv file and store here
        datalist = []

        # open a csv file
        csvfile = open(filename, "r")

        # create csv reader
        reader = csv.reader(csvfile)

        # skip header
        next(reader)

        # Add csv rows to list
        for rows in reader:
            datalist.append(rows)
        return datalist
